import openpyxl
from geopy import Nominatim
import os
import django
import errno
import time

os.chdir('/home/chase/DissertationProjects/ArboModel/ArboModelData/Databases/WeatherDB/Weather_Data_Scraping/')
print(os.system('pwd'))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'ArboModelData.settings')

django.setup()
from Databases.WeatherDB.models import Countries, Dates, Counties, LatitudeLongitude, WeatherStations, WeatherSensors, SensorType


def  get_geo_point_meta(latitude, longitude):
    time.sleep(1)
    geolocator = Nominatim(user_agent="geoapiExercises")

    location = geolocator.reverse(str(latitude)+","+str(longitude))

    return location.raw['address']


def dms2dd(latitude, longitude):
    def convert_to_degrees(lat_or_lon):
        out_degrees = float(lat_or_lon[0]) + float(lat_or_lon[1])/60 + float(lat_or_lon[2])/(60*60)
        return out_degrees
    lat_spt = latitude.strip().split()
    lon_spt = longitude.strip().split()

    out_lat = round(convert_to_degrees(lat_spt),3)
    out_lon = round(convert_to_degrees(lon_spt) * -1,3)

    return (out_lat, out_lon)


def parse_alert_station_summary(file):
    """

    :param file:
    :return: Dictionary of station information
    """
    try:
        wb_obj = openpyxl.load_workbook(file)
        sheet = wb_obj.active

    except FileNotFoundError:
        raise FileNotFoundError(errno.ENOENT, os.strerror(errno.ENOENT), file)

    data_dict_output = {}
    for row in sheet.iter_rows(min_row=3):
        row_data = []
        for cell in row:
            row_data.append(cell.value)


        data_dict_output.update(
            {
            row_data[0].strip().lower().replace('.',''):{
                'station_type': row_data[1].strip().lower(),
                'date_installed': row_data[2],
                'old_station_id': (lambda x: int(x) if ( x != None) else None)(row_data[3]),
                'new_station_id': (lambda x: int(x) if ( x != None) else None)(row_data[4]),
                'old_rain_id':(lambda x: int(x) if ( x != None) else None)(row_data[5]),
                'new_rain_id': (lambda x: int(x) if ( x != None) else None)(row_data[6]),
                'old_water_level_id': (lambda x: int(x) if ( x != None) else None)(row_data[7]),
                'new_water_level_id':(lambda x: int(x) if ( x != None) else None)(row_data[8])
                }
            }
        )

    return data_dict_output



def parse_alert1_all_sensors(file, alert_summary_dict, start_row):
    """

    :param file:
    :param alert_summary_dict:
    :return:
    """
    try:
        wb_obj = openpyxl.load_workbook(file)
        sheet = wb_obj.active

    except FileNotFoundError:
        raise FileNotFoundError(errno.ENOENT, os.strerror(errno.ENOENT), file)


    for row in sheet.iter_rows(min_row=start_row):
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        lat_lon_tmp = dms2dd(row_data[4], row_data[5])
        data_dict = {
        'station_name' : row_data[0].strip().lower().replace('.',''),
        'device_id' : row_data[1],
        'dev_type' : row_data[2],
        'installed' : row_data[3],
        'lat_lon' : lat_lon_tmp,
        'location_data' : get_geo_point_meta(lat_lon_tmp[0], lat_lon_tmp[1]),
        'elevation' : int(row_data[6]),
        'station_description' : row_data[7]
        }

        print(alert_summary_dict[data_dict['station_name']])
        print(data_dict)
        date_station, date_station_created = Dates.objects.using('WeatherDB').get_or_create(date=alert_summary_dict[data_dict['station_name']]['date_installed'])

        if date_station_created:
            date_station.save()

        date_sensor, date_sensor_created = Dates.objects.using('WeatherDB').get_or_create(date=data_dict['installed'])

        if date_sensor_created:
            date_sensor.save()

        country, country_created =  Countries.objects.using('WeatherDB').get_or_create(country = "united states",
                                                                    country_code = "us"
                                                                    )
        if country_created:
            country.save()

        try:
            county, county_created = Counties.objects.using('WeatherDB').get_or_create(
                county = data_dict['location_data']['county'].strip().lower())
        except KeyError:
            county, county_created = Counties.objects.using('WeatherDB').get_or_create(
                county='unknown')
        if county_created:
            county.save()


        lat_lon, lat_lon_created = LatitudeLongitude.objects.using('WeatherDB').get_or_create(latitude=data_dict['lat_lon'][0],
                                                                           longitude=data_dict['lat_lon'][1]
                                                                           )
        if lat_lon_created:
            lat_lon.save()

        weather_station, weather_station_created = WeatherStations.objects.using('WeatherDB').get_or_create(name=data_dict['station_name'],
                                                                                         lat_lon=lat_lon,
                                                                                         country=country,
                                                                                         county=county,
                                                                                         station_id=alert_summary_dict[data_dict['station_name']]['new_station_id'],
                                                                                         old_station_id=alert_summary_dict[data_dict['station_name']]['old_station_id'],
                                                                                         elevation=data_dict['elevation'],
                                                                                         station_loc_description=data_dict['station_description'],
                                                                                         installed_date_station=date_station,
                                                                                         new_rain_id=alert_summary_dict[data_dict['station_name']]['new_rain_id'],
                                                                                         old_rain_id=alert_summary_dict[data_dict['station_name']]['old_rain_id'],
                                                                                         new_water_level_id=alert_summary_dict[data_dict['station_name']]['new_water_level_id'],
                                                                                         old_water_level_id=alert_summary_dict[data_dict['station_name']]['old_water_level_id']
                                                                                         )
        if weather_station_created:
            weather_station.save()

        sensor_type, sensor_type_created = SensorType.objects.using('WeatherDB').get_or_create(sensor_type=data_dict['dev_type'])
        if sensor_type_created:
            sensor_type.save()

        weather_sensors, weather_sensors_created = WeatherSensors.objects.using('WeatherDB').get_or_create(station=weather_station,
                                                                                        installed_date_sensor=date_sensor,
                                                                                        sensor_id=data_dict['device_id'],
                                                                                        sensor_type=sensor_type
                                                                                        )
        if weather_sensors_created:
            weather_sensors.save()




if __name__ == '__main__':
    alert_summary_dict = parse_alert_station_summary('/home/chase/DissertationProjects/ArboModel/ArboModelData/Databases/WeatherDB/Weather_Data_Scraping/ALERT_Data/ALERT2_Numbering_Summary.xlsx')
    parse_alert1_all_sensors('/home/chase/DissertationProjects/ArboModel/ArboModelData/Databases/WeatherDB/Weather_Data_Scraping/ALERT_Data/ALERT_sensors_all_by_ID.xlsx',
                             alert_summary_dict,
                             2)